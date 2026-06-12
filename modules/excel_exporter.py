# modules/excel_exporter.py
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from modules.config_manager import load_settings
from datetime import datetime

# ── Palette ───────────────────────────────────────────────────────────────────
_NAVY    = "0B1E35"
_NAVY_MD = "1A3A60"
_GOLD    = "F5A800"
_WHITE   = "FFFFFF"
_STRIPE  = "EDF4FD"
_META_BG = "F4F8FD"
_META_FG = "7090A8"
_GOOD_BG = "D4EDDA"
_GOOD_FG = "0E5C23"
_WARN_BG = "FFF3CD"
_WARN_FG = "7C5C00"
_BAD_BG  = "FAD7DA"
_BAD_FG  = "7C1422"
_BORDER_CLR = "C8D8E8"

# ── Shared style atoms ────────────────────────────────────────────────────────
_T     = Side(style="thin",   color=_BORDER_CLR)
_GT    = Side(style="thin",   color=_GOLD)
_GM    = Side(style="medium", color=_GOLD)

_B     = Border(left=_T, right=_T, top=_T,  bottom=_T)
_B_HDR = Border(left=_T, right=_T, top=_GT, bottom=_GM)
_B_SUM = Border(left=_T, right=_T, top=_GM, bottom=_T)

_CTR = Alignment(horizontal="center", vertical="center", readingOrder=2)
_RGT = Alignment(horizontal="right",  vertical="center", readingOrder=2)


# ── Low-level helpers ─────────────────────────────────────────────────────────

def _fill(color):
    return PatternFill("solid", fgColor=color)

def _font(color, bold=False, size=11):
    return Font(color=color, bold=bold, size=size, name="Calibri")

def _status_color(value, good_thr, warn_thr, higher_is_better=True):
    if higher_is_better:
        if value >= good_thr: return _GOOD_BG, _GOOD_FG
        if value >= warn_thr: return _WARN_BG, _WARN_FG
    else:
        if value <= good_thr: return _GOOD_BG, _GOOD_FG
        if value <= warn_thr: return _WARN_BG, _WARN_FG
    return _BAD_BG, _BAD_FG

def _bonus_color(total):
    if total >= 1500: return _GOOD_BG, _GOOD_FG
    if total >= 800:  return _WARN_BG, _WARN_FG
    return _BAD_BG, _BAD_FG

def _cell(ws, row, col, value, bg=None, fg=None, fmt=None, bold=False, align=None):
    stripe = _STRIPE if row % 2 == 0 else _WHITE
    c = ws.cell(row, col, value)
    c.fill      = _fill(bg if bg is not None else stripe)
    c.font      = _font(fg or _NAVY, bold=bold)
    c.alignment = align or _CTR
    c.border    = _B
    if fmt:
        c.number_format = fmt
    return c

def _bonus_cell(ws, row, col, amount, row_bg):
    bg = _GOOD_BG if amount > 0 else row_bg
    fg = _GOOD_FG if amount > 0 else _META_FG
    _cell(ws, row, col, amount, bg=bg, fg=fg, fmt="#,##0 ₪")

def _summary_row(ws, row, ncols, cells):
    for ci in range(1, ncols + 1):
        c = ws.cell(row, ci)
        c.fill = _fill(_NAVY); c.font = _font(_GOLD, bold=True)
        c.border = _B_SUM;     c.alignment = _CTR
    for entry in cells:
        col, val = entry[0], entry[1]
        fmt = entry[2] if len(entry) > 2 else None
        c = ws.cell(row, col, val)
        c.fill = _fill(_NAVY); c.font = _font(_GOLD, bold=True)
        c.border = _B_SUM
        c.alignment = _RGT if col == 1 else _CTR
        if fmt: c.number_format = fmt
    ws.row_dimensions[row].height = 26

def _header_row(ws, cols):
    for ci, label in enumerate(cols, 1):
        c = ws.cell(4, ci, label)
        c.fill = _fill(_NAVY); c.font = _font(_GOLD, bold=True, size=11)
        c.alignment = _CTR; c.border = _B_HDR
    ws.row_dimensions[4].height = 26

def _brand_header(ws, title, ncols, meta=""):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    c = ws.cell(1, 1, "⚡  Volta Solar  |  מוקד תיאומים")
    c.fill = _fill(_NAVY); c.font = Font(name="Calibri", bold=True, size=13, color=_GOLD)
    c.alignment = _CTR; ws.row_dimensions[1].height = 36

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    c = ws.cell(2, 1, title)
    c.fill = _fill(_NAVY_MD); c.font = Font(name="Calibri", bold=True, size=16, color=_WHITE)
    c.alignment = _CTR; ws.row_dimensions[2].height = 32

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=ncols)
    c = ws.cell(3, 1, meta)
    c.fill = _fill(_META_BG); c.font = Font(name="Calibri", size=9, color=_META_FG)
    c.alignment = _CTR; ws.row_dimensions[3].height = 18

def _init_sheet(ws, title, ncols, meta="", tab_color=None):
    ws.sheet_view.rightToLeft   = True
    ws.sheet_view.showGridLines = False
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    _brand_header(ws, title, ncols, meta)
    return 5  # data starts at row 5

def _autofit(ws, overrides=None):
    overrides = overrides or {}
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        if letter in overrides:
            ws.column_dimensions[letter].width = overrides[letter]
        else:
            w = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[letter].width = min(max(w + 4, 10), 40)

def _print_setup(ws):
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
    ws.print_options.horizontalCentered = True


# ── Weekly KPI ────────────────────────────────────────────────────────────────

def export_weekly_kpi(kpi_data: list, filepath: str) -> None:
    s   = load_settings()["bonus_thresholds"]
    now = datetime.now().strftime("%d/%m/%Y %H:%M")

    mph_good  = s["meetings_per_hour_tier_a"]
    mph_warn  = mph_good * 0.75
    occ_good  = s["occupancy_tier_a_pct"] / 100
    occ_warn  = s["occupancy_tier_b_pct"] / 100
    idle_good = s["idle_tier_a_pct"] / 100
    idle_warn = s["idle_tier_b_pct"] / 100

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "KPI שבועי"

    COLS  = ["נציג", "שעות עבודה", "תיאומים", "תיאומים/שעה",
             "תעסוקה %", "סרק %", "פניקס", "שיחות נענו"]
    ncols = len(COLS)
    ds    = _init_sheet(ws, "דוח KPI שבועי", ncols,
                        meta=f"הופק: {now}  |  נציגים: {len(kpi_data)}",
                        tab_color=_GOLD)
    _header_row(ws, COLS)
    ws.freeze_panes = "A5"

    for ri, a in enumerate(kpi_data, ds):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws.row_dimensions[ri].height = 22
        _cell(ws, ri, 1, a["name"],             bg=stripe, align=_RGT, bold=True)
        _cell(ws, ri, 2, round(a["hours"], 1),  bg=stripe, fmt="0.0")
        _cell(ws, ri, 3, a["meetings"],          bg=stripe)

        mph = round(a["meetings_per_hour"], 2)
        bg, fg = _status_color(mph, mph_good, mph_warn, True)
        _cell(ws, ri, 4, mph, bg=bg, fg=fg, bold=True, fmt="0.00")

        occ = a["occupancy_pct"]
        bg, fg = _status_color(occ, occ_good, occ_warn, True)
        _cell(ws, ri, 5, occ, bg=bg, fg=fg, fmt="0.0%")

        idl = a["idle_pct"]
        bg, fg = _status_color(idl, idle_good, idle_warn, False)
        _cell(ws, ri, 6, idl, bg=bg, fg=fg, fmt="0.00%")

        _cell(ws, ri, 7, a["phoenix"],                bg=stripe)
        _cell(ws, ri, 8, a.get("answered_calls", 0),  bg=stripe)

    n  = len(kpi_data)
    tm = sum(a["meetings"]               for a in kpi_data)
    th = sum(a["hours"]                  for a in kpi_data)
    ta = sum(a.get("answered_calls", 0)  for a in kpi_data)
    tp = sum(a["phoenix"]                for a in kpi_data)
    cmph     = round(tm / th, 2) if th else 0
    avg_occ  = sum(a["occupancy_pct"] for a in kpi_data) / n if n else 0
    avg_idle = sum(a["idle_pct"]      for a in kpi_data) / n if n else 0

    sr = n + ds
    _summary_row(ws, sr, ncols, [
        (1, "סיכום מוקד"),
        (2, round(th, 1), "0.0"),
        (3, tm),
        (4, cmph, "0.00"),
        (7, tp),
        (8, ta),
    ])
    for col, val, fmt, bg, fg in [
        (5, avg_occ,  "0.0%",  *_status_color(avg_occ,  occ_good,  occ_warn,  True)),
        (6, avg_idle, "0.00%", *_status_color(avg_idle, idle_good, idle_warn, False)),
    ]:
        c = ws.cell(sr, col, val)
        c.fill = _fill(bg); c.font = _font(fg, bold=True)
        c.border = _B_SUM; c.alignment = _CTR; c.number_format = fmt

    _autofit(ws, {"A": 22, "B": 16, "C": 14, "D": 16,
                  "E": 14, "F": 14, "G": 12, "H": 16})
    _print_setup(ws)
    wb.save(filepath)


# ── Monthly Bonus — center-wide workbook ──────────────────────────────────────

def export_monthly_bonus(bonus_data: list, billing: dict,
                         month_label: str, filepath: str,
                         kpi_data: list = None,
                         manager_bonus: float = 0,
                         manager_name: str = "טל סנדר",
                         center_meets: bool = False) -> None:

    s              = load_settings()["bonus_thresholds"]
    occ_good       = s["occupancy_tier_a_pct"] / 100
    occ_warn       = s["occupancy_tier_b_pct"] / 100
    idle_good      = s["idle_tier_a_pct"] / 100
    idle_warn      = s["idle_tier_b_pct"] / 100
    mph_good       = s["meetings_per_hour_tier_a"]
    mph_warn       = mph_good * 0.75
    rate_a         = s["meetings_per_hour_tier_a_rate"]   # 5 ₪
    rate_b         = s["meetings_per_hour_tier_b_rate"]   # 4 ₪
    center_rate    = s.get("center_target_bonus_per_meeting", 1)  # +1 ₪ if center meets
    ph_emp_rate    = s["phoenix_employee_rate"]    # 50 ₪ — agent payment (לתשלום)
    ph_client_rate = s["phoenix_client_rate"]      # 100 ₪ — billing detail (פירוט)

    now          = datetime.now().strftime("%d/%m/%Y %H:%M")
    kpi_map      = {k["name"]: k for k in (kpi_data or [])}
    agents_total = sum(b["total"] for b in bonus_data)
    grand_total  = agents_total + manager_bonus

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ── 1. לתשלום — agents + manager, agent phoenix @ 50 ₪ ──────────────────
    ws = wb.create_sheet("לתשלום")
    n  = 3
    ds = _init_sheet(ws, f"בונוסים לתשלום  |  {month_label}", n,
                     meta=f"הופק: {now}  |  נציגים: {len(bonus_data)}  |  סה\"כ ₪{grand_total:,.0f}",
                     tab_color=_NAVY)
    _header_row(ws, ["שם", "מספר עובד / תפקיד", "בונוס לתשלום (₪)"])
    ws.freeze_panes = "A5"

    for ri, b in enumerate(bonus_data, ds):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws.row_dimensions[ri].height = 22
        _cell(ws, ri, 1, b["name"],        bg=stripe, align=_RGT, bold=True)
        _cell(ws, ri, 2, b["employee_id"], bg=stripe)
        bg, fg = _bonus_color(b["total"])
        _cell(ws, ri, 3, b["total"], bg=bg, fg=fg, bold=True, fmt="#,##0 ₪")

    mgr_ri = len(bonus_data) + ds
    ws.row_dimensions[mgr_ri].height = 22
    _cell(ws, mgr_ri, 1, manager_name, bg="E8F0FE", align=_RGT, bold=True)
    _cell(ws, mgr_ri, 2, "מנהל מוקד",  bg="E8F0FE")
    mbg, mfg = _bonus_color(manager_bonus)
    _cell(ws, mgr_ri, 3, manager_bonus, bg=mbg, fg=mfg, bold=True, fmt="#,##0 ₪")

    _summary_row(ws, mgr_ri + 1, n, [
        (1, 'סה"כ לתשלום'),
        (3, grand_total, "#,##0 ₪"),
    ])
    _autofit(ws, {"A": 22, "B": 20, "C": 22})
    _print_setup(ws)

    # ── 2. פירוט בונוסים — 15 cols, phoenix @ 100 ₪ ─────────────────────────
    ws2 = wb.create_sheet("פירוט בונוסים")
    COLS2 = [
        "שם נציג", "שעות", "תיאומים", "תיאומים/שעה",
        "עמלת תיאומים", "תעסוקה %", "בונוס תעסוקה",
        "שיחות סרק", "% סרק", "בונוס סרק",
        "ציון משוב", "בונוס משוב",
        f"עסקת פניקס ({ph_client_rate}₪)", "בונוס ליעד צוותי",
        'סה"כ',
    ]
    n2  = len(COLS2)
    cm  = "✅ מוקד עמד ביעד" if center_meets else "❌ מוקד לא עמד ביעד"
    ds2 = _init_sheet(ws2, f"פירוט בונוסים  |  {month_label}", n2,
                      meta=f"הופק: {now}  |  {cm}  |  נציגים: {len(bonus_data)}",
                      tab_color="1B5FAA")
    _header_row(ws2, COLS2)
    ws2.freeze_panes = "A5"

    detail_totals = []
    for ri, b in enumerate(bonus_data, ds2):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws2.row_dimensions[ri].height = 22
        k    = kpi_map.get(b["name"], {})
        mph  = k.get("meetings_per_hour", 0)
        occ  = k.get("occupancy_pct", 0)
        idl  = k.get("idle_pct", 0)
        mtg  = k.get("meetings", 0)
        ph   = k.get("phoenix", 0)
        fs   = k.get("feedback_score")
        hrs  = round(k.get("hours", 0), 1)

        base_r    = rate_a if mph >= mph_good else rate_b
        mtg_base  = mtg * base_r
        ctr_bonus = mtg * center_rate if center_meets else 0
        ph_val    = ph * ph_client_rate
        row_total = mtg_base + ctr_bonus + b["occupancy_bonus"] + b["idle_bonus"] + b["feedback_bonus"] + ph_val
        detail_totals.append(row_total)

        _cell(ws2, ri,  1, b["name"],          bg=stripe, align=_RGT, bold=True)
        _cell(ws2, ri,  2, hrs,                 bg=stripe, fmt="0.0")
        _cell(ws2, ri,  3, mtg,                 bg=stripe)
        bg, fg = _status_color(mph, mph_good, mph_warn, True)
        _cell(ws2, ri,  4, round(mph, 2),       bg=bg, fg=fg, fmt="0.00")
        _bonus_cell(ws2, ri,  5, mtg_base,      stripe)
        bg, fg = _status_color(occ, occ_good, occ_warn, True)
        _cell(ws2, ri,  6, occ,                 bg=bg, fg=fg, fmt="0.0%")
        _bonus_cell(ws2, ri,  7, b["occupancy_bonus"], stripe)
        _cell(ws2, ri,  8, k.get("idle_calls", 0), bg=stripe)
        bg, fg = _status_color(idl, idle_good, idle_warn, False)
        _cell(ws2, ri,  9, idl,                 bg=bg, fg=fg, fmt="0.00%")
        _bonus_cell(ws2, ri, 10, b["idle_bonus"], stripe)
        fb_bg = _GOOD_BG if fs and fs >= 8.5 else _WARN_BG if fs and fs >= 8.0 else stripe
        _cell(ws2, ri, 11, fs if fs is not None else "—",
              bg=fb_bg, fmt="0.0" if fs else None)
        _bonus_cell(ws2, ri, 12, b["feedback_bonus"], stripe)
        _bonus_cell(ws2, ri, 13, ph_val,         stripe)
        _bonus_cell(ws2, ri, 14, ctr_bonus,      stripe)
        bg, fg = _bonus_color(row_total)
        _cell(ws2, ri, 15, row_total, bg=bg, fg=fg, bold=True, fmt="#,##0 ₪")

    _summary_row(ws2, len(bonus_data) + ds2, n2, [
        (1,  'סה"כ'),
        (3,  sum(k.get("meetings", 0) for k in kpi_map.values())),
        (15, sum(detail_totals), "#,##0 ₪"),
    ])
    _autofit(ws2, {"A": 20, "B": 10, "C": 12, "D": 14, "E": 16,
                   "F": 12, "G": 16, "H": 14, "I": 10, "J": 14,
                   "K": 12, "L": 14, "M": 18, "N": 18, "O": 14})
    _print_setup(ws2)

    # ── 3. סיכום מוקד ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("סיכום מוקד")
    n3  = 4
    ds3 = _init_sheet(ws3, f"סיכום מוקד  |  {month_label}", n3,
                      meta=f"הופק: {now}", tab_color="217346")
    _header_row(ws3, ["מדד", "ערך", "יעד", "סטטוס"])

    summary_rows = []
    if kpi_data:
        tm  = sum(k["meetings"]      for k in kpi_data)
        th  = sum(k["hours"]         for k in kpi_data)
        tph = sum(k["phoenix"]       for k in kpi_data)
        ao  = sum(k["occupancy_pct"] for k in kpi_data) / len(kpi_data)
        ai  = sum(k["idle_pct"]      for k in kpi_data) / len(kpi_data)
        cmph = tm / th if th else 0
        summary_rows = [
            ("קצב מוקד (תיאומים/שעה)", round(cmph, 2),       f"≥{mph_good}",    cmph >= mph_good),
            ('סה"כ תיאומים',           tm,                    "—",               None),
            ('סה"כ שעות עבודה',        round(th, 1),          "—",               None),
            ("ממוצע תעסוקה",            f"{ao*100:.1f}%",     f"≥{occ_good*100:.0f}%", ao >= occ_good),
            ("ממוצע סרק",               f"{ai*100:.2f}%",     f"≤{idle_good*100:.0f}%", ai <= idle_good),
            ('סה"כ פניקס (עסקאות)',    tph,                   "—",               None),
        ]
    summary_rows += [
        ('סה"כ בונוסים נציגים (₪)',           agents_total,  "—", None),
        (f"בונוס מנהל — {manager_name} (₪)",  manager_bonus, "—", None),
        ('סה"כ לתשלום (₪)',                    grand_total,   "—", None),
    ]
    for ri, (label, val, target, met) in enumerate(summary_rows, ds3):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws3.row_dimensions[ri].height = 22
        _cell(ws3, ri, 1, label,  bg=stripe, align=_RGT, bold=True)
        _cell(ws3, ri, 2, val,    bg=stripe)
        _cell(ws3, ri, 3, target, bg=stripe)
        if met is True:
            _cell(ws3, ri, 4, "✅ עמד ביעד", bg=_GOOD_BG, fg=_GOOD_FG, bold=True)
        elif met is False:
            _cell(ws3, ri, 4, "❌ לא עמד",   bg=_BAD_BG,  fg=_BAD_FG,  bold=True)
        else:
            _cell(ws3, ri, 4, "—", bg=stripe)
    _autofit(ws3, {"A": 34, "B": 18, "C": 14, "D": 18})
    _print_setup(ws3)

    # ── 4. חיוב ללקוח — hours + phoenix @ 100 ₪ ─────────────────────────────
    ws4 = wb.create_sheet("חיוב ללקוח")
    n4  = 2
    ph_count   = billing.get("phoenix_count", 0)
    ph_billing = billing.get("phoenix_billing", 0)
    ds4 = _init_sheet(ws4, f"חיוב ללקוח  |  {month_label}", n4,
                      meta=f"הופק: {now}  |  פניקס: {ph_count} עסקאות",
                      tab_color=_GOLD)
    _header_row(ws4, ["נציג", "שעות עבודה"])
    ws4.freeze_panes = "A5"

    for ri, (name, hours) in enumerate(billing["hours_by_agent"].items(), ds4):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws4.row_dimensions[ri].height = 22
        _cell(ws4, ri, 1, name,             bg=stripe, align=_RGT, bold=True)
        _cell(ws4, ri, 2, round(hours, 1),  bg=stripe, fmt="0.0")

    sr4 = len(billing["hours_by_agent"]) + ds4
    _summary_row(ws4, sr4, n4, [
        (1, 'סה"כ שעות'),
        (2, round(billing["total_hours"], 1), "0.0"),
    ])
    _summary_row(ws4, sr4 + 1, n4, [
        (1, f"פניקס ({ph_count} עסקאות × {ph_client_rate}₪)"),
        (2, ph_billing, "#,##0 ₪"),
    ])
    _autofit(ws4, {"A": 28, "B": 18})
    _print_setup(ws4)

    wb.save(filepath)


# ── Per-agent bonus file ───────────────────────────────────────────────────────

def export_agent_bonus(kpi: dict, bonus: dict,
                       month_label: str, filepath: str,
                       center_meets: bool = False) -> None:

    s              = load_settings()["bonus_thresholds"]
    occ_good       = s["occupancy_tier_a_pct"] / 100
    occ_warn       = s["occupancy_tier_b_pct"] / 100
    idle_good      = s["idle_tier_a_pct"] / 100
    idle_warn      = s["idle_tier_b_pct"] / 100
    mph_good       = s["meetings_per_hour_tier_a"]
    mph_warn       = mph_good * 0.75
    rate_a         = s["meetings_per_hour_tier_a_rate"]
    rate_b         = s["meetings_per_hour_tier_b_rate"]
    center_rate    = s.get("center_target_bonus_per_meeting", 1)
    ph_emp_rate    = s["phoenix_employee_rate"]
    now            = datetime.now().strftime("%d/%m/%Y %H:%M")

    mph      = kpi.get("meetings_per_hour", 0)
    meetings = kpi.get("meetings", 0)
    occ      = kpi.get("occupancy_pct", 0)
    idl      = kpi.get("idle_pct", 0)
    fs       = kpi.get("feedback_score")
    ph_count = kpi.get("phoenix", 0)
    base_r   = rate_a if mph >= mph_good else rate_b
    mtg_base = meetings * base_r
    ctr_b    = meetings * center_rate if center_meets else 0

    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = "בונוס אישי"
    ncols = 4  # מדד | ביצועים | יעד | בונוס
    ds    = _init_sheet(ws, f"דוח בונוס אישי  —  {kpi['name']}", ncols,
                        meta=f"{month_label}  |  הופק: {now}",
                        tab_color=_GOLD)
    _header_row(ws, ["מדד", "ביצועים", "יעד", "בונוס (₪)"])

    # each entry: (label, perf_value, perf_fmt, target_str, perf_bg, perf_fg, bonus_amount_or_None)
    rows = [
        ("שעות עבודה",
         round(kpi.get("hours", 0), 1), "0.0",
         "—", None, None, None),

        ("תיאומים",
         meetings, None,
         "—", None, None, None),

        ("תיאומים לשעה",
         round(mph, 2), "0.00",
         f"≥ {mph_good}",
         *_status_color(mph, mph_good, mph_warn, True), None),

        ("עמלת תיאומים",
         f"{meetings} × {base_r}₪", None,
         f"שער {'A' if base_r == rate_a else 'B'}",
         _GOOD_BG, _GOOD_FG, mtg_base),

        ("בונוס ליעד צוותי",
         "✅ עמד" if center_meets else "❌ לא עמד", None,
         f"+{center_rate}₪ לתיאום",
         _GOOD_BG if center_meets else _BAD_BG,
         _GOOD_FG if center_meets else _BAD_FG,
         ctr_b),

        ("אחוז תעסוקה",
         occ, "0.0%",
         f"≥ {occ_good*100:.0f}%",
         *_status_color(occ, occ_good, occ_warn, True),
         bonus["occupancy_bonus"]),

        ("שיחות סרק",
         kpi.get("idle_calls", 0), None,
         "—", None, None, None),

        ("אחוז סרק",
         idl, "0.00%",
         f"≤ {idle_good*100:.0f}%",
         *_status_color(idl, idle_good, idle_warn, False),
         bonus["idle_bonus"]),

        ("ציון משוב",
         fs if fs is not None else "—", "0.0" if fs else None,
         "≥ 8.0",
         _GOOD_BG if fs and fs >= 8.5 else _WARN_BG if fs and fs >= 8.0 else None, None,
         bonus["feedback_bonus"]),

        ("עסקת פניקס",
         f"{ph_count} עסקאות", None,
         f"{ph_emp_rate}₪ לעסקה",
         _GOOD_BG if ph_count > 0 else None, _GOOD_FG if ph_count > 0 else None,
         bonus["phoenix_bonus"]),
    ]

    for ri, (label, perf, pfmt, target, p_bg, p_fg, bval) in enumerate(rows, ds):
        stripe = _STRIPE if ri % 2 == 0 else _WHITE
        ws.row_dimensions[ri].height = 22
        _cell(ws, ri, 1, label,  bg=stripe, align=_RGT, bold=True)
        _cell(ws, ri, 2, perf,   bg=p_bg or stripe, fg=p_fg, fmt=pfmt)
        _cell(ws, ri, 3, target, bg=stripe, fg=_META_FG)
        if bval is not None:
            _bonus_cell(ws, ri, 4, bval, stripe)
        else:
            _cell(ws, ri, 4, "—", bg=stripe, fg=_META_FG)

    total     = bonus["total"]
    total_row = ds + len(rows)
    ws.row_dimensions[total_row].height = 28
    bg, fg = _bonus_color(total)
    _summary_row(ws, total_row, ncols, [
        (1, 'סה"כ בונוס לתשלום'),
        (4, total, "#,##0 ₪"),
    ])
    ws.cell(total_row, 4).fill = _fill(bg)
    ws.cell(total_row, 4).font = _font(fg, bold=True, size=14)

    _autofit(ws, {"A": 26, "B": 18, "C": 22, "D": 16})
    _print_setup(ws)
    wb.save(filepath)
