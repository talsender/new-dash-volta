import pandas as pd
import datetime, re


def _parse_hours(val) -> float:
    if isinstance(val, datetime.time):
        return val.hour + val.minute / 60 + val.second / 3600
    if isinstance(val, datetime.timedelta):
        return val.total_seconds() / 3600
    if isinstance(val, (int, float)):
        if pd.isna(val):
            return 0.0
        if 0 < val < 1:
            return val * 24  # Excel fraction of day
        return float(val)
    s = str(val).strip()
    if ':' in s:
        parts = s.split(':')
        try:
            return int(parts[0]) + int(parts[1]) / 60
        except (ValueError, IndexError):
            return 0.0
    try:
        v = float(s)
        return v * 24 if 0 < v < 1 else v
    except (ValueError, TypeError):
        return 0.0


def parse_attendance(filepath: str) -> pd.DataFrame:
    with pd.ExcelFile(filepath, engine='openpyxl') as xl:
        sheet = next((s for s in xl.sheet_names if 'וולטה' in s), xl.sheet_names[0])
        raw = xl.parse(sheet, header=None)
        header_row = next(
            (i for i, row in raw.iterrows()
             if any('מספר' in str(v) and 'עובד' in str(v) for v in row.values)),
            None
        )
        if header_row is None:
            raise KeyError("לא נמצאה שורת כותרות עם 'מספר עובד'")
        df = xl.parse(sheet, header=header_row)
    df.columns = [str(c).strip() for c in df.columns]
    df = df[pd.to_numeric(df['מספר עובד'], errors='coerce').notna()].copy()
    df['מספר עובד'] = pd.to_numeric(df['מספר עובד'], errors='coerce').astype('Int64')
    # sum per-shift columns (סה"כ 1, סה"כ 2, ...) — more reliable than סה"כ כללי
    shift_cols = [c for c in df.columns if re.match(r'סה"כ\s*\d+', c)]
    if shift_cols:
        df['סה"כ כללי'] = sum(df[c].apply(_parse_hours) for c in shift_cols)
    else:
        hours_col = next((c for c in df.columns if 'סה"כ' in c or 'כללי' in c), None)
        if hours_col is None:
            raise KeyError(f"לא נמצאה עמודת שעות. עמודות: {list(df.columns)}")
        df['סה"כ כללי'] = df[hours_col].apply(_parse_hours)
    return df


def parse_voicenter(filepath: str) -> pd.DataFrame:
    raw = None
    for enc in ('utf-16', 'utf-8', 'windows-1255'):
        try:
            tables = pd.read_html(filepath, encoding=enc, header=None)
            if tables:
                raw = tables[0]
                break
        except Exception:
            continue
    if raw is None:
        raise KeyError("לא ניתן לקרוא את קובץ Voicenter")

    raw = raw.astype(str).apply(lambda col: col.str.strip())
    col_names = [str(c).strip() for c in raw.columns]

    # case 1: pandas already detected headers (e.g. <th> tags)
    if any('משתמש' in c for c in col_names):
        raw.columns = col_names
        df = raw
    else:
        # case 2: all columns are numeric — find header row in values
        header_row = next(
            (i for i, row in raw.iterrows() if any('משתמש' in str(v) for v in row.values)),
            None
        )
        if header_row is None:
            raise KeyError(f"לא נמצאה שורת כותרות עם 'משתמש'. עמודות: {col_names}")
        raw.columns = [str(v).strip() for v in raw.iloc[header_row]]
        df = raw.iloc[header_row + 1:].reset_index(drop=True)

    user_col = next((c for c in df.columns if 'משתמש' in c), None)
    df = df.rename(columns={user_col: 'משתמש'})
    df = df[df['משתמש'].notna() & ~df['משתמש'].astype(str).str.startswith('סה"כ') & (df['משתמש'] != 'nan')]

    occ_col = next((c for c in df.columns if 'תעסוקה' in c), None)
    if occ_col is None:
        raise KeyError(f"לא נמצאה עמודת תעסוקה. עמודות: {list(df.columns)}")
    df = df.rename(columns={occ_col: 'אחוז תעסוקה נטו'})
    df['אחוז תעסוקה נטו'] = (df['אחוז תעסוקה נטו'].astype(str)
                              .str.replace('%', '', regex=False).str.strip()
                              .pipe(pd.to_numeric, errors='coerce') / 100)

    answered_col = next((c for c in df.columns if c == 'נענו' or ('נענו' in c and 'לא' not in c)), None)
    if answered_col:
        df = df.rename(columns={answered_col: 'נענו'})
    df['נענו'] = pd.to_numeric(df.get('נענו', 0), errors='coerce').fillna(0).astype(int)

    # Total incoming calls (כניסות) — distinct from answered (נענו)
    _total_candidates = ['כניסות', 'שיחות נכנסות', 'שיחות']
    total_col = next(
        (c for c in df.columns
         if str(c).strip() in _total_candidates
         or any(t in str(c) for t in _total_candidates)),
        None,
    )
    if total_col and total_col != answered_col:
        df = df.rename(columns={total_col: 'כניסות'})
        df['כניסות'] = pd.to_numeric(df['כניסות'], errors='coerce').fillna(0).astype(int)
    else:
        df['כניסות'] = df['נענו']  # fallback: same as answered

    return df.reset_index(drop=True)


def parse_feedback(filepath: str) -> dict:
    import openpyxl
    scores = {}
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sname in wb.sheetnames:
        ws = wb[sname]
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            label = str(row[0]).strip() if row[0] is not None else ""
            if label and ('ציון' in label):
                # find first numeric value in the row (column index >= 1)
                for cell in row[1:]:
                    try:
                        score = float(cell)
                        scores[sname.strip()] = score
                        break
                    except (TypeError, ValueError):
                        continue
                break
    return scores
