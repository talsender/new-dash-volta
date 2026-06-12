# tests/test_excel_exporter.py
import tempfile, os
import openpyxl
from modules.excel_exporter import export_weekly_kpi, export_monthly_bonus

def _kpi():
    return [{"name": "טום", "hours": 80.0, "meetings": 80, "meetings_per_hour": 1.0,
             "occupancy_pct": 0.35, "idle_pct": 0.01, "phoenix": 3}]

def _bonus():
    return [{"name": "טום", "employee_id": 98804, "meetings_bonus": 480,
             "occupancy_bonus": 300, "idle_bonus": 150,
             "feedback_bonus": 150, "phoenix_bonus": 150, "total": 1230}]

def _billing():
    return {"hours_by_agent": {"טום": 80.0}, "total_hours": 80.0,
            "phoenix_count": 3, "phoenix_billing": 300}

def test_export_weekly_creates_sheet():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        export_weekly_kpi(_kpi(), path)
        wb = openpyxl.load_workbook(path)
        assert 'KPI שבועי' in wb.sheetnames
    finally:
        os.unlink(path)

def test_export_weekly_has_agent_name():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        export_weekly_kpi(_kpi(), path)
        wb = openpyxl.load_workbook(path)
        ws = wb['KPI שבועי']
        names = [ws.cell(r, 1).value for r in range(2, ws.max_row + 1)]
        assert 'טום' in names
    finally:
        os.unlink(path)

def test_export_monthly_has_three_sheets():
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
        path = f.name
    try:
        export_monthly_bonus(_bonus(), _billing(), "יוני 2026", path)
        wb = openpyxl.load_workbook(path)
        assert {'לתשלום', 'פירוט בונוסים', 'חיוב ללקוח'}.issubset(set(wb.sheetnames))
    finally:
        os.unlink(path)
